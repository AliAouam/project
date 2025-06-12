import os
import io
from datetime import datetime
from typing import List, Optional

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Path, status, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import FileResponse
from pydantic import BaseModel

import motor.motor_asyncio
from passlib.context import CryptContext
from PIL import Image
import torch
import torch.nn as nn
import torch.nn.functional as F
from torchvision import transforms
from bson import ObjectId

from fpdf import FPDF
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
import json

# ------------------------
# FastAPI + CORS
# ------------------------
app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ------------------------
# MongoDB
# ------------------------
MONGODB_URI = "mongodb://localhost:27017"
client = motor.motor_asyncio.AsyncIOMotorClient(MONGODB_URI)
db = client.retinal_annotation

# ------------------------
# Static uploads + CORS for images
# ------------------------
BASE_DIR = os.path.dirname(__file__)
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

class CORSMiddlewareStatic(StaticFiles):
    async def __call__(self, scope, receive, send):
        async def send_wrapper(message):
            if message["type"] == "http.response.start":
                headers = dict(message["headers"])
                headers[b'access-control-allow-origin'] = b'*'
                message["headers"] = list(headers.items())
            await send(message)
        await super().__call__(scope, receive, send_wrapper)

app.mount("/uploads", CORSMiddlewareStatic(directory=UPLOAD_DIR), name="uploads")

# ------------------------
# Password hashing
# ------------------------
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ------------------------
# Pydantic models
# ------------------------
class UserIn(BaseModel):
    email: str
    name: str
    password: str
    role: str

class UserOut(BaseModel):
    id: str
    email: str
    name: str
    role: str

class ImageOut(BaseModel):
    id: str
    patientId: str
    patientName: str
    filename: str
    url: str
    uploadedAt: datetime
    uploadedBy: Optional[str] = None

class AnnotationIn(BaseModel):
    imageId: str
    x: float
    y: float
    width: float
    height: float
    type: str
    severity: str
    color: str
    created_by: Optional[str] = None

class AnnotationOut(AnnotationIn):
    id: str
    createdAt: datetime

class AIPrediction(BaseModel):
    label: str
    confidence: float

class ClassificationIn(BaseModel):
    patientId: str
    patientName: str
    imagePath: str
    manual_label: str
    stage: Optional[int] = None
    other_disease: Optional[str] = None
    ai_prediction: Optional[AIPrediction] = None
    annotations: List[dict]
    comparison: str
    exported_at: str

class ClassificationOut(ClassificationIn):
    id: str
    created_at: datetime

class LogEntry(BaseModel):
    id: Optional[str]
    action: str
    entity: str
    entity_id: str
    user: Optional[str]
    details: Optional[dict] = None
    created_at: datetime

# ------------------------
# IA Model
# ------------------------
class RetinalCNN(nn.Module):
    def __init__(self, num_classes: int = 5):
        super().__init__()
        self.conv1 = nn.Conv2d(3, 32, 3, padding=1)
        self.conv2 = nn.Conv2d(32, 64, 3, padding=1)
        self.pool = nn.MaxPool2d(2, 2)
        self.fc1 = nn.Linear(64 * 56 * 56, 256)
        self.fc2 = nn.Linear(256, num_classes)

    def forward(self, x):
        x = self.pool(torch.relu(self.conv1(x)))
        x = self.pool(torch.relu(self.conv2(x)))
        x = x.view(x.size(0), -1)
        x = torch.relu(self.fc1(x))
        return self.fc2(x)

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model = RetinalCNN(5).to(device)
model.load_state_dict(torch.load("models/retinal_model.pth", map_location=device))
model.eval()

CLASS_NAMES = ["Mild", "Moderate", "No_DR", "Proliferate_DR", "Severe"]
transform = transforms.Compose([
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
])

def predict_image(data: bytes) -> dict:
    try:
        img = Image.open(io.BytesIO(data)).convert("RGB")
        img = img.resize((224, 224))
        t = transform(img).unsqueeze(0).to(device)
        out = model(t)
        probs = F.softmax(out, dim=1)
        _, p = torch.max(out, 1)
        return {
            "label": CLASS_NAMES[p.item()],
            "confidence": float(probs[0, p.item()] * 100),
        }
    except Exception as e:
        print("Erreur lors de la prédiction IA:", str(e))
        return {
            "label": "—",
            "confidence": 0.0
        }

# ------------------------
# LOGS UTILS
# ------------------------
async def add_log(action, entity, entity_id, user=None, details=None):
    entry = {
        "action": action,
        "entity": entity,
        "entity_id": entity_id,
        "user": user,
        "details": details,
        "created_at": datetime.utcnow(),
    }
    await db.logs.insert_one(entry)

# ------------------------
# User endpoints
# ------------------------
@app.post("/api/users", response_model=UserOut)
async def create_user(u: UserIn):
    if await db.users.find_one({"email": u.email}):
        raise HTTPException(400, "Email already registered")
    hashed = pwd_context.hash(u.password)
    doc = {"email": u.email, "name": u.name, "role": u.role, "hashed_password": hashed}
    res = await db.users.insert_one(doc)
    await add_log("create", "user", str(res.inserted_id), user=u.email)
    return UserOut(id=str(res.inserted_id), email=u.email, name=u.name, role=u.role)

@app.post("/api/login", response_model=UserOut)
async def login(form_data: OAuth2PasswordRequestForm = Depends()):
    user = await db.users.find_one({"email": form_data.username})
    if not user or not pwd_context.verify(form_data.password, user.get("hashed_password", "")):
        raise HTTPException(401, "Invalid credentials")
    return UserOut(id=str(user["_id"]), email=user["email"], name=user["name"], role=user["role"])

@app.get("/api/users", response_model=List[UserOut])
async def list_users():
    docs = await db.users.find().to_list(100)
    return [UserOut(id=str(d["_id"]), email=d["email"], name=d["name"], role=d["role"]) for d in docs]

@app.delete("/api/users/{user_id}")
async def delete_user(user_id: str):
    res = await db.users.delete_one({"_id": ObjectId(user_id)})
    await add_log("delete", "user", user_id)
    return {"deleted": bool(res.deleted_count)}

# ------------------------
# Image endpoints
# ------------------------
@app.post("/api/images", response_model=ImageOut)
async def upload_image(
    patientId: str = Form(...),
    patientName: str = Form(...),
    image: UploadFile = File(...),
    uploadedBy: Optional[str] = Form(None)
):
    if not image.content_type.startswith("image/"):
        raise HTTPException(400, "File must be an image")
    dest = os.path.join(UPLOAD_DIR, image.filename)
    with open(dest, "wb") as f:
        f.write(await image.read())
    public_url = f"/uploads/{image.filename}"
    doc = {
        "patientId": patientId,
        "patientName": patientName,
        "filename": image.filename,
        "url": public_url,
        "uploadedAt": datetime.utcnow(),
        "uploadedBy": uploadedBy,
    }
    res = await db.images.insert_one(doc)
    await add_log("upload", "image", str(res.inserted_id))
    return ImageOut(id=str(res.inserted_id), **doc)

@app.get("/api/images", response_model=List[ImageOut])
async def list_images(uploaded_by: Optional[str] = Query(None)):
    query = {}
    if uploaded_by:
        query["uploadedBy"] = uploaded_by
    docs = await db.images.find(query).to_list(100)
    return [ImageOut(id=str(d["_id"]), **d) for d in docs]

@app.delete("/api/images/{image_id}", status_code=status.HTTP_200_OK)
async def delete_image(image_id: str):
    img_res = await db.images.delete_one({"_id": ObjectId(image_id)})
    ann_res = await db.annotations.delete_many({"imageId": image_id})
    await add_log("delete", "image", image_id)
    return {
        "deleted_images": img_res.deleted_count,
        "deleted_annotations": ann_res.deleted_count
    }

# ------------------------
# Annotation endpoints
# ------------------------
@app.get("/api/annotations", response_model=List[AnnotationOut])
async def get_all_annotations():
    docs = await db.annotations.find().to_list(10000)
    return [AnnotationOut(id=str(d["_id"]), **d) for d in docs]

@app.get("/api/annotations/{imageId}", response_model=List[AnnotationOut])
async def get_annotations(
    imageId: str = Path(..., description="Image ID"),
    created_by: Optional[str] = Query(None)
):
    query = {"imageId": imageId}
    if created_by:
        query["created_by"] = created_by
    docs = await db.annotations.find(query).to_list(100)
    return [AnnotationOut(id=str(d["_id"]), **d) for d in docs]

@app.post("/api/annotations", response_model=AnnotationOut)
async def create_annotation(ann: AnnotationIn):
    d = ann.dict()
    d["createdAt"] = datetime.utcnow()
    res = await db.annotations.insert_one(d)
    await add_log("create", "annotation", str(res.inserted_id), user=d.get("created_by"))
    return AnnotationOut(id=str(res.inserted_id), **d)

@app.delete("/api/annotations/{annotation_id}")
async def delete_annotation(annotation_id: str = Path(..., description="Annotation ID")):
    res = await db.annotations.delete_one({"_id": ObjectId(annotation_id)})
    if res.deleted_count:
        await add_log("delete", "annotation", annotation_id)
        return {"deleted": True}
    raise HTTPException(404, "Annotation not found")

# ------------------------
# AI & Classification endpoints
# ------------------------
@app.post("/api/predict", response_model=AIPrediction)
async def predict(file: UploadFile = File(...)):
    data = await file.read()
    result = predict_image(data)
    return AIPrediction(**result)

@app.post("/api/classifications", response_model=ClassificationOut)
async def save_classification(classif: ClassificationIn):
    d = classif.dict()
    d["created_at"] = datetime.utcnow()
    res = await db.classifications.insert_one(d)
    await add_log("export", "classification", str(res.inserted_id))
    return ClassificationOut(id=str(res.inserted_id), **d)

@app.get("/api/classifications", response_model=List[ClassificationOut])
async def get_classifications():
    docs = await db.classifications.find().to_list(100)
    return [ClassificationOut(id=str(d["_id"]), **d) for d in docs]

# ------------------------
# LOGS endpoints
# ------------------------
@app.get("/api/logs", response_model=List[LogEntry])
async def get_logs(limit: int = 100):
    docs = await db.logs.find().sort("created_at", -1).to_list(limit)
    return [
        LogEntry(
            id=str(d["_id"]),
            action=d["action"],
            entity=d["entity"],
            entity_id=d["entity_id"],
            user=d.get("user"),
            details=d.get("details"),
            created_at=d["created_at"]
        )
        for d in docs
    ]

# ------------------------
# Export PDF endpoint
# ------------------------
@app.get("/api/export/pdf/{image_id}")
async def export_annotations_pdf(image_id: str):
    image_doc = await db.images.find_one({"_id": ObjectId(image_id)})
    if not image_doc:
        raise HTTPException(404, "Image not found")

    anns = await db.annotations.find({"imageId": image_id}).to_list(100)
    image_path = os.path.join(UPLOAD_DIR, image_doc["filename"])

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Annotation Report", ln=1, align='C')

    try:
        pdf.image(image_path, x=10, y=30, w=120)
    except:
        pass

    pdf.ln(60)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f'Patient: {image_doc["patientName"]}  (ID: {image_doc["patientId"]})', ln=1)
    pdf.cell(0, 10, f'Image file: {image_doc["filename"]}', ln=1)
    pdf.cell(0, 10, f'Uploaded: {image_doc["uploadedAt"].strftime("%Y-%m-%d %H:%M:%S")}', ln=1)

    pdf.ln(8)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Annotations:", ln=1)
    pdf.set_font("Arial", "", 12)
    if anns:
        for a in anns:
            user = a.get("created_by", "-")
            date_str = a["createdAt"].strftime("%Y-%m-%d %H:%M") if isinstance(a["createdAt"], datetime) else str(a["createdAt"])
            pdf.cell(0, 8, f'- {a["type"].capitalize()} | Stage: {a["severity"]} | User: {user} | Date: {date_str}', ln=1)
    else:
        pdf.cell(0, 8, "Aucune annotation.", ln=1)

    out_path = os.path.join(UPLOAD_DIR, f"annotation_report_{image_id}.pdf")
    pdf.output(out_path)
    await add_log("export", "pdf", image_id)
    return FileResponse(out_path, media_type="application/pdf", filename=f"annotation_report_{image_id}.pdf")

# ------------------------
# Export Excel endpoint
# ------------------------
@app.get("/api/export/excel")
async def export_logs_excel():
    logs = await db.logs.find().sort("created_at", -1).to_list(1000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    ws.append(["Action", "Entity", "Entity ID", "User", "Details", "Created At"])

    for l in logs:
        created = l["created_at"].strftime("%Y-%m-%d %H:%M:%S") if isinstance(l["created_at"], datetime) else str(l["created_at"])
        ws.append([
            l.get("action"),
            l.get("entity"),
            l.get("entity_id"),
            l.get("user"),
            json.dumps(l.get("details", {})),
            created
        ])

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        out_path = tmp.name

    await add_log("export", "excel", "logs")
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="logs.xlsx",
    )

# ------------------------
# Run
# ------------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
